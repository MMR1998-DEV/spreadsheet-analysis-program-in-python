import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

total_products_from_suppliers = {}
total_value_per_suppliers = {}
product_under_10 = {}
#print(product_list.max_row)
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventiry = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inv_total_price = product_list.cell(product_row, 5)
    # calculation products per supplier
    if supplier_name in total_products_from_suppliers:
        current = total_products_from_suppliers.get(supplier_name)
        total_products_from_suppliers[supplier_name] = current + 1
    else:        
        total_products_from_suppliers[supplier_name] = 1

    # calculation total value of inventory

    if supplier_name in total_value_per_suppliers:
        current_total_value = total_value_per_suppliers.get(supplier_name)
        total_value_per_suppliers[supplier_name] = current_total_value + inventiry * price

    else:
        total_value_per_suppliers[supplier_name] = inventiry * price

    
    #login products is under 10 units
    if inventiry < 10:
        product_under_10[int(product_number)] = int(inventiry)

    #total inventory price

    inv_total_price.value = inventiry * price

print(total_products_from_suppliers)
print(total_value_per_suppliers)
print(product_under_10)

inv_file.save("inv_new.xlsx")
